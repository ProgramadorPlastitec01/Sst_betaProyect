from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, View, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy, reverse
from django.db import transaction, models
from django.db.models import Q, Count
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.contenttypes.models import ContentType
from django.utils import timezone
from django.contrib.auth import get_user_model

from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models import (
    Area, InspectionSchedule, InspectionSignature,
    ExtinguisherInspection, ExtinguisherItem,
    FirstAidInspection, FirstAidItem, FirstAidSignature,
    ProcessInspection, ProcessSignature, ProcessCheckItem,
    StorageInspection, StorageCheckItem, StorageSignature,
    ForkliftInspection, ForkliftCheckItem, ForkliftSignature,
    InspectionEvidence
)
from .forms import (
    InspectionScheduleForm, InspectionUpdateForm,
    ExtinguisherInspectionForm, ExtinguisherItemFormSet, ExtinguisherItemForm,
    FirstAidInspectionForm, FirstAidItemFormSet, FirstAidItemForm,
    ProcessInspectionForm, ProcessItemFormSet, ProcessCheckItemForm,
    StorageInspectionForm, StorageItemFormSet, StorageCheckItemForm,
    ForkliftInspectionForm, ForkliftItemFormSet, ForkliftCheckItemForm
)
from notifications.models import NotificationGroup, Notification
from roles.mixins import RolePermissionRequiredMixin

# --- Mixin to provide form user context ---
class InspectionFormUserMixin:
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

def handle_pending_evidences(request, instance, prefix=None):
    from django.contrib.contenttypes.models import ContentType
    from .models import InspectionEvidence
    content_type = ContentType.objects.get_for_model(instance)
    key = f'pending_evidences_{content_type.id}'
    if prefix:
        key = f'{key}_{prefix}'
    
    files = request.FILES.getlist(key)
    for f in files:
        InspectionEvidence.objects.create(
            content_type=content_type,
            object_id=instance.pk,
            image=f,
            description='',
            uploaded_by=request.user
        )

# --- Mixin to provide evidence context ---
class EvidenceMixin:
    """
    Mixin to provide ContentType IDs for evidence management in templates.
    Handles mapping from Inspection models to their Item models.
    """
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Determine the primary model for this view
        model_to_use = getattr(self, 'model', None)
        if not model_to_use and hasattr(self, 'object') and self.object:
            model_to_use = self.object.__class__
        elif not model_to_use and hasattr(self, 'get_queryset'):
            model_to_use = self.get_queryset().model
            
        if model_to_use:
            # Map inspection models to their respective item models
            inspection_item_map = [
                (ExtinguisherInspection, ExtinguisherItem),
                (FirstAidInspection, FirstAidItem),
                (ProcessInspection, ProcessCheckItem),
                (StorageInspection, StorageCheckItem),
                (ForkliftInspection, ForkliftCheckItem),
            ]
            
            # All item model types (for reverse lookup)
            all_item_models = [item for _, item in inspection_item_map]
            
            # Legacy/Primary content type for the object being handled
            context['inspection_content_type_id'] = ContentType.objects.get_for_model(model_to_use).id
            
            # Find the item model if the current model is an inspection
            for insp, item in inspection_item_map:
                if model_to_use == insp:
                    context['item_content_type_id'] = ContentType.objects.get_for_model(item).id
                    break
            
            # Reverse lookup: if the current model IS an item model, set item_content_type_id directly
            if model_to_use in all_item_models:
                context['item_content_type_id'] = ContentType.objects.get_for_model(model_to_use).id

        return context

# --- Mixin to provide matrix context to any view ---
class MatrixContextMixin:
    def get_context_data(self, **kwargs):
        from datetime import date
        context = super().get_context_data(**kwargs)
        current_year = date.today().year

        # Pass unique values for filters with pre-calculated selection state
        # Default to current year if not specified in GET (initial load)
        # If specified but empty (?year=), it means 'All'
        get_year = self.request.GET.get('year')
        selected_year = str(current_year) if get_year is None else get_year
        
        selected_area = self.request.GET.get('area', '')
        selected_type = self.request.GET.get('inspection_type', '')
        
        years_qs = list(InspectionSchedule.objects.values_list('scheduled_date__year', flat=True).distinct())
        if current_year not in years_qs:
            years_qs.append(current_year)
        years_qs.sort(reverse=True)
        context['years_data'] = [{'val': str(y), 'selected': str(y) == selected_year} for y in years_qs if y is not None]
        
        # FIX: Use area name instead of ID for dropdown
        areas_qs = InspectionSchedule.objects.values_list('area__id', 'area__name').distinct().order_by('area__name')
        try:
             sel_area_id = int(selected_area) if selected_area else None
        except ValueError:
             sel_area_id = None
        context['areas_data'] = [{'id': a[0], 'name': a[1], 'selected': a[0] == sel_area_id} for a in areas_qs]
        
        # Display year for the table header
        context['year_display'] = selected_year if selected_year else "Histórico Completo"
        
        context['statuses'] = InspectionSchedule.STATUS_CHOICES
        
        # FORM FOR MODAL - FIXING THE EMPTY FORM ISSUE
        context['form'] = InspectionScheduleForm()
        
        # Base Queryset for Matrix (respect perms and filters)
        # Re-assigned later after permission checks and filter logic in point 2.
        pass

        # 1. Permission Mapping
        user = self.request.user
        inspection_modules = [
            {'name': 'Extintores', 'key': 'extinguisher', 'patterns': ['extintor'], 'model': ExtinguisherInspection},
            {'name': 'Botiquines', 'key': 'first_aid', 'patterns': ['botiquin'], 'model': FirstAidInspection},
            {'name': 'Instalaciones de Proceso', 'key': 'process', 'patterns': ['proceso', 'instalacion'], 'model': ProcessInspection},
            {'name': 'Almacenamiento', 'key': 'storage', 'patterns': ['almacen', 'storage'], 'model': StorageInspection},
            {'name': 'Montacargas', 'key': 'forklift', 'patterns': ['montacarga'], 'model': ForkliftInspection},
        ]
        
        allowed_keys = []
        allowed_types = []
        model_mapping = {}
        type_to_key = {} # Map 'Extintores' -> 'extinguisher'

        from django.db.models import Q
        type_filter_q = Q(pk__in=[])

        for mod in inspection_modules:
            if user.has_perm_custom(mod['key'], 'view'):
                allowed_keys.append(mod['key'])
                allowed_types.append(mod['name'])
                model_mapping[mod['name']] = mod['model']
                type_to_key[mod['name']] = mod['key']
                # Filter Q for database
                t_q = Q()
                for p in mod['patterns']:
                    t_q |= Q(inspection_type__icontains=p)
                type_filter_q |= t_q

        found_types = list(InspectionSchedule.objects.filter(type_filter_q).values_list('inspection_type', flat=True).distinct())
        all_types = sorted(list(set(allowed_types + found_types)))

        # New context for filter dropdown
        context['types_data'] = [{'val': t, 'selected': t == selected_type} for t in all_types]
        
        # Actually filter the processing list if requested
        if selected_type:
            all_types = [t for t in all_types if t == selected_type]

        # 2. Base Queryset (respect perms and filters)
        matrix_qs = InspectionSchedule.objects.filter(type_filter_q)
        
        if selected_year: 
            matrix_qs = matrix_qs.filter(scheduled_date__year=selected_year)
            
        if selected_area: 
            matrix_qs = matrix_qs.filter(area__id=selected_area)

        if selected_type:
            matrix_qs = matrix_qs.filter(inspection_type=selected_type)

        evidence_mapping = {
            'Extintores': 'Inspección de extintores (R-RH-SST-019)',
            'Botiquines': 'Inspección de Botiquines (R-RH-SST-020)',
            'Instalaciones de Proceso': 'Inspección de Instalaciones (R-RH-SST-030)',
            'Almacenamiento': 'Inspección Almacenamiento (R-RH-SST-031)',
            'Montacargas': 'Inspección de Montacargas (R-RH-SST-022)',
        }

        matrix = []
        months_range = range(1, 13)
        months_names = ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEPT', 'OCT', 'NOV', 'DIC']

        for t in all_types:
            type_qs = matrix_qs.filter(inspection_type=t)
            row_cells = []
            total_p = 0
            total_e = 0
            actual_model = model_mapping.get(t)
            
            # Get responsible from the first schedule item found for this type
            first_item = type_qs.first()
            responsible_name = first_item.responsible.get_full_name() if first_item and first_item.responsible else "Equipo SST"

            for m in months_range:
                p_items = type_qs.filter(scheduled_date__month=m)
                p_count = p_items.count()
                # Count scheduled items that are effectively done
                e_scheduled = p_items.filter(status__in=['Realizada', 'Cerrada', 'Cerrada con Hallazgos', 'Cumple', 'No Cumple']).count()
                
                e_count = e_scheduled
                
                if actual_model:
                    actual_qs = actual_model.objects.filter(inspection_date__month=m)
                    if selected_year: actual_qs = actual_qs.filter(inspection_date__year=selected_year)
                    if selected_area: actual_qs = actual_qs.filter(area_id=selected_area)
                    
                    # Exclude follow-ups from metrics
                    if hasattr(actual_model, 'parent_inspection'):
                         actual_qs = actual_qs.filter(parent_inspection__isnull=True)
                         
                    unlinked_count = actual_qs.filter(schedule_item__isnull=True).count()
                    e_count += unlinked_count

                total_p += p_count
                total_e += e_count
                
                row_cells.append({
                    'm': m,
                    'p': p_count if p_count > 0 else '',
                    'e': e_count if e_count > 0 else '',
                    'e_scheduled': e_scheduled
                })
            
            compliance = (total_e / total_p * 100) if total_p > 0 else (100 if total_e > 0 else 0)
            matrix.append({
                'type': t,
                'cells': row_cells,
                'total_p': total_p,
                'total_e': total_e,
                'compliance': round(compliance),
                'responsible': responsible_name,
                'evidence': evidence_mapping.get(t, 'SST Record')
            })
        context['matrix'] = matrix
        context['months_names'] = months_names
        
        # Bottom Summary Table
        total_p_g = sum(item['total_p'] for item in matrix)
        total_e_g = sum(item['total_e'] for item in matrix)

        # ---------------------------------------------------------
        # MODAL CHART LOGIC: Global Compliance (All Modules)
        # ---------------------------------------------------------
        modal_monthly_stats = []
        modal_max_comp = 0
        
        # Initialize month totals (Ignoring Area Filter)
        months_totals = {m: {'p': 0, 'e': 0} for m in range(1, 13)}

        for t in all_types:
            # Base schedule query for this type (Ignoring Area)
            type_qs_global = InspectionSchedule.objects.filter(inspection_type=t)
            if selected_year: type_qs_global = type_qs_global.filter(year=selected_year)
            
            actual_model = model_mapping.get(t)

            for m in range(1, 13):
                p_items = type_qs_global.filter(scheduled_date__month=m)
                # P = Total Scheduled
                p_count = p_items.count()
                # E = Items marked as done in schedule
                e_scheduled = p_items.filter(status__in=['Realizada', 'Cerrada', 'Cerrada con Hallazgos', 'Cumple', 'No Cumple']).count()
                
                e_count = e_scheduled
                
                # Plus Unlinked Executions
                if actual_model:
                    actual_qs = actual_model.objects.filter(inspection_date__month=m)
                    if selected_year: actual_qs = actual_qs.filter(inspection_date__year=selected_year)
                    # Skip Area Filter here for global modal
                    
                    if hasattr(actual_model, 'parent_inspection'):
                        actual_qs = actual_qs.filter(parent_inspection__isnull=True)
                    
                    unlinked_count = actual_qs.filter(schedule_item__isnull=True).count()
                    e_count += unlinked_count
                
                months_totals[m]['p'] += p_count
                months_totals[m]['e'] += e_count

        # Prepare final stats and find max for scaling
        raw_final = []
        for m in range(1, 13):
            p = months_totals[m]['p']
            e = months_totals[m]['e']
            comp = (e / p * 100) if p > 0 else (100 if e > 0 else 0)
            if comp > modal_max_comp:
                modal_max_comp = comp
            raw_final.append({'m': m, 'p': p, 'e': e, 'compliance': comp})

        # Scaling and context prep
        modal_y_max = 100 if modal_max_comp <= 100 else modal_max_comp
        modal_usable_height = 300

        for idx, item in enumerate(raw_final):
            h_px = (item['compliance'] / modal_y_max * modal_usable_height) if modal_y_max > 0 else 0
            modal_monthly_stats.append({
                'month_name': months_names[idx],
                'compliance': round(item['compliance']),
                'height_px': int(h_px),
                'p': item['p'],
                'e': item['e']
            })

        context['modal_monthly_stats'] = modal_monthly_stats
        context['modal_y_max_label'] = round(modal_y_max)
        context['modal_y_mid_label'] = round(modal_y_max / 2)

        # Global totals for cards
        context['total_p_global'] = total_p_g
        context['total_e_global'] = total_e_g
        context['global_compliance'] = round((total_e_g / total_p_g * 100) if total_p_g > 0 else (100 if total_e_g > 0 else 0))
        
        return context

# --- Mixin to provide scheduled inspections to module list views ---
class ScheduledInspectionsMixin:
    """
    Mixin to add pending scheduled inspections to any inspection module list view.
    Automatically filters by inspection type based on the module.
    """
    # Map of module types to their inspection_type keywords
    INSPECTION_TYPE_MAP = {
        'extinguisher': 'Extintores',
        'first_aid': 'Botiquin',
        'process': 'Instalaciones de Proceso',
        'storage': 'Almacenamiento',
        'forklift': 'Montacargas',
    }
    
    # Override this in the view to specify which module type
    inspection_module_type = None
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.inspection_module_type and self.inspection_module_type in self.INSPECTION_TYPE_MAP:
            keyword = self.INSPECTION_TYPE_MAP[self.inspection_module_type]
            
            qs = InspectionSchedule.objects.filter(
                inspection_type__icontains=keyword
            )
            
            schedule_id = self.request.GET.get('schedule_id')
            if schedule_id:
                qs = qs.filter(id=schedule_id)
            else:
                qs = qs.filter(status='Programada')
                
            # Enrich items with timeline logic
            today = timezone.now().date()
            enhanced_items = []
            for item in qs.select_related('responsible').order_by('scheduled_date'):
                # item.is_overdue is now a model property, do not overwrite
                item.is_due_soon = today <= item.scheduled_date <= (today + timedelta(days=5))
                # Allow execution ONLY if the scheduled date is today or in the past (overdue)
                item.can_execute = item.scheduled_date <= today
                enhanced_items.append(item)
                
            context['scheduled_inspections'] = enhanced_items
        else:
            context['scheduled_inspections'] = []
        
        return context

# 0. Core Schedule Views
class InspectionListView(LoginRequiredMixin, MatrixContextMixin, ListView):
    model = InspectionSchedule
    template_name = 'inspections/inspection_list.html'
    context_object_name = 'schedule'

    def get_queryset(self):
        from datetime import date
        qs = super().get_queryset()
        user = self.request.user
        
        # Permission filter
        from django.db.models import Q
        inspection_modules = [
            {'key': 'extinguisher', 'patterns': ['extintor']},
            {'key': 'first_aid', 'patterns': ['botiquin']},
            {'key': 'process', 'patterns': ['proceso', 'instalacion']},
            {'key': 'storage', 'patterns': ['almacen', 'storage']},
            {'key': 'forklift', 'patterns': ['montacarga']},
        ]
        type_filter_q = Q(pk__in=[])
        for mod in inspection_modules:
            if user.has_perm_custom(mod['key'], 'view'):
                t_q = Q()
                for p in mod['patterns']:
                    t_q |= Q(inspection_type__icontains=p)
                type_filter_q |= t_q
        
        qs = qs.filter(type_filter_q)
        
        get_year = self.request.GET.get('year')
        selected_year = str(date.today().year) if get_year is None else get_year
        
        area = self.request.GET.get('area')
        
        if selected_year: 
            qs = qs.filter(scheduled_date__year=selected_year)
        if area: 
            qs = qs.filter(area__id=area)
        
        selected_type = self.request.GET.get('inspection_type')
        if selected_type:
            qs = qs.filter(inspection_type=selected_type)
            
        return qs.order_by('-scheduled_date')
    
    def get_context_data(self, **kwargs):
        from datetime import date
        context = super().get_context_data(**kwargs)
        current_year = date.today().year
        
        # Get filter parameters
        get_year = self.request.GET.get('year')
        year_filter = str(current_year) if get_year is None else get_year
        area_filter = self.request.GET.get('area', '')
        type_filter = self.request.GET.get('inspection_type', '')
        
        # Consolidate all inspections from all modules
        all_inspections = []
        
        # Helper function to add inspections to list
        def add_inspections(model, inspection_type, detail_url_name):
            qs = model.objects.all()
            
            # Filter Only Initial Inspections (Hide Follow-ups)
            if hasattr(model, 'parent_inspection'):
                 qs = qs.filter(parent_inspection__isnull=True)
            
            # Apply filters
            if year_filter:
                qs = qs.filter(inspection_date__year=int(year_filter))
            if area_filter:
                qs = qs.filter(area__id=area_filter)
            
            for insp in qs:
                # Use the model method to get total follow-ups count
                follow_ups_count = insp.get_total_follow_ups_count() if hasattr(insp, 'get_total_follow_ups_count') else 0
                
                all_inspections.append({
                    'id': insp.pk,
                    'date': insp.inspection_date,
                    'year': insp.inspection_date.year,
                    'area': insp.area,
                    'type': inspection_type,
                    'inspector': insp.inspector.get_full_name() if insp.inspector else 'N/A',
                    'status': getattr(insp, 'status', insp.general_status),
                    'detail_url': reverse(detail_url_name, args=[insp.pk]),
                    'schedule_linked': insp.schedule_item is not None,
                    'follow_ups_count': follow_ups_count
                })
        
        # Add inspections from each module ONLY if user has permission
        permission_map = [
            (ExtinguisherInspection, 'extinguisher', 'Extintores', 'extinguisher_detail'),
            (FirstAidInspection, 'first_aid', 'Botiquines', 'first_aid_detail'),
            (ProcessInspection, 'process', 'Instalaciones de Proceso', 'process_detail'),
            (StorageInspection, 'storage', 'Almacenamiento', 'storage_detail'),
            (ForkliftInspection, 'forklift', 'Montacargas', 'forklift_detail'),
        ]
        
        allowed_models_for_filters = []

        for model_cls, perm_key, label, url_name in permission_map:
            if self.request.user.has_perm_custom(perm_key, 'view'):
                # Apply type filter if selected
                if not type_filter or type_filter == label:
                    add_inspections(model_cls, label, url_name)
                    allowed_models_for_filters.append(model_cls)
        
        # Sort by date descending
        all_inspections.sort(key=lambda x: x['date'], reverse=True)
        
        # Limit mainly displayed lists to 10
        context['all_inspections'] = all_inspections
        context['latest_inspections'] = all_inspections[:10]
        
        # ── Lógica de Alertas de Activos (Dashboard) ─────────────────
        from gestion_activos.models import Asset
        # Pre-fetching de detalles para optimizar el cálculo del property 'estado_actual'
        assets_qs = Asset.objects.select_related('asset_type').prefetch_related('extintor_detail', 'montacargas_detail').all()
        
        vencidos_count = 0
        proximos_count = 0
        criticos_count = 0
        
        for asset in assets_qs:
            st = asset.estado_actual
            if st in ('VENCIDO', 'MANTENIMIENTO_VENCIDO'):
                vencidos_count += 1
            elif st in ('PROXIMO_A_VENCER', 'PROXIMO_MANTENIMIENTO'):
                proximos_count += 1
            
            # Se consideran críticos aquellos fuera de servicio (activo=False)
            if not asset.activo:
                criticos_count += 1
        
        context['asset_alerts'] = {
            'vencidos': vencidos_count,
            'proximos': proximos_count,
            'criticos': criticos_count,
            'total': vencidos_count + proximos_count + criticos_count,
            'has_alerts': (vencidos_count + proximos_count + criticos_count) > 0
        }
        # ─────────────────────────────────────────────────────────────

        # Determine full and latest schedule
        # context['schedule'] is already set by ListView
        # Assuming get_queryset returns the full list filtered by year/area
        context['full_schedule'] = context['schedule']
        context['latest_schedule'] = context['schedule'][:10]
        
        # Get unique years and areas from actual inspections for filters
        all_years = set()
        all_areas = set()
        
        for model in allowed_models_for_filters:
            years = model.objects.values_list('inspection_date__year', flat=True).distinct()
            all_years.update(years)
            areas = model.objects.values_list('area__id', 'area__name').distinct()
            all_areas.update(areas)
        
        # Also include years/areas from schedule
        schedule_years = InspectionSchedule.objects.values_list('scheduled_date__year', flat=True).distinct()
        all_years.update(schedule_years)
        schedule_areas = InspectionSchedule.objects.values_list('area__id', 'area__name').distinct()
        all_areas.update(schedule_areas)
        
        # Ensure current year is always present
        all_years.add(current_year)
        
        # Update filter data with actual inspection data
        context['years_data'] = [{'val': str(y), 'selected': str(y) == year_filter} for y in sorted(list(all_years), reverse=True)]
        # Sort by area name (index 1 in tuple)
        sorted_areas = sorted(list(all_areas), key=lambda x: x[1])
        context['areas_data'] = [{'id': a[0], 'name': a[1], 'selected': str(a[0]) == str(area_filter)} for a in sorted_areas]
        
        return context

class InspectionCreateView(LoginRequiredMixin, CreateView):
    model = InspectionSchedule
    form_class = InspectionScheduleForm
    template_name = 'inspections/inspection_form.html'
    success_url = reverse_lazy('inspection_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        base = self.object
        
        # Robust year extraction
        try:
            target_year = int(form.cleaned_data.get('year'))
        except (ValueError, TypeError):
            target_year = base.scheduled_date.year
            
        current_date = base.scheduled_date
        
        # Robust frequency mapping
        freq_map = {
            'Mensual': 1,
            'Bimestral': 2,
            'Trimestral': 3,
            'Cuatrimestral': 4,
            'Semestral': 6,
            'Anual': 12
        }
        
        months_to_add = freq_map.get(base.frequency, 0)
        created_count = 0
        
        if months_to_add > 0:
            next_date = current_date + relativedelta(months=months_to_add)
            # Generate only within the target year
            while next_date.year == target_year:
                if not InspectionSchedule.objects.filter(
                    area=base.area,
                    inspection_type=base.inspection_type,
                    scheduled_date=next_date
                ).exists():
                    InspectionSchedule.objects.create(
                        year=target_year,
                        area=base.area,
                        inspection_type=base.inspection_type,
                        frequency=base.frequency,
                        scheduled_date=next_date,
                        responsible=base.responsible,
                        status='Programada',
                        observations=base.observations
                    )
                    created_count += 1
                next_date += relativedelta(months=months_to_add)

        if created_count > 0:
            messages.success(self.request, f'Se han generado {created_count} programaciones adicionales automáticas para {target_year}.')
        else:
            messages.success(self.request, f'Programación guardada exitosamente.')
            
        return response

class InspectionUpdateView(LoginRequiredMixin, UpdateView):
    model = InspectionSchedule
    form_class = InspectionUpdateForm
    template_name = 'inspections/inspection_form.html'
    success_url = reverse_lazy('inspection_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Programación actualizada correctamente')
        return response

class InspectionDeleteView(LoginRequiredMixin, DeleteView):
    model = InspectionSchedule
    template_name = 'inspections/inspection_confirm_delete.html'
    success_url = reverse_lazy('inspection_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Programación eliminada exitosamente')
        return super().delete(request, *args, **kwargs)

# Helper to link inspections
def link_to_schedule(request, obj):
    s_id = request.GET.get('schedule_item') or request.POST.get('schedule_item')
    if s_id:
        try:
            item = InspectionSchedule.objects.get(pk=s_id)
            obj.schedule_item = item
            obj.save()
            
            # Mark as done
            item.status = 'Realizada'
            item.save()
            
            # Generate next recurrence
            next_schedule = item.generate_next_schedule()
            if next_schedule:
                messages.info(request, f"📅 Nueva programación generada automáticamente para el {next_schedule.scheduled_date.strftime('%d/%m/%Y')}")
                
        except Exception as e:
            print(f"Error linking schedule: {e}")

# --- Mixin for Formsets ---
class FormsetMixin:
    formset_class = None
    initial_items = []

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['items'] = self.formset_class(self.request.POST, instance=self.object)
        else:
            # Check if existing object has no items (e.g. failed creation or manual deletion)
            object_has_items = False
            if getattr(self, 'object', None) and hasattr(self.object, 'items'):
                object_has_items = self.object.items.exists()

            if (not getattr(self, 'object', None) or (getattr(self, 'object', None) and not object_has_items)) and self.initial_items:
                initial = [{'question': q} for q in self.initial_items]
                context['items'] = self.formset_class(instance=self.object, initial=initial)
                context['items'].extra = len(initial)
            else:
                context['items'] = self.formset_class(instance=self.object)
        
        # SI ES UN SEGUIMIENTO (tiene parent_inspection), forzamos extra=0
        if getattr(self, 'object', None) and hasattr(self.object, 'parent_inspection') and self.object.parent_inspection:
            context['items'].extra = 0
            
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        items = context['items']
        with transaction.atomic():
            self.object = form.save()
            handle_pending_evidences(self.request, self.object)
            link_to_schedule(self.request, self.object)
            if items.is_valid():
                # Iterate directly over forms to keep prefix always in sync
                for item_form in items.forms:
                    if item_form in items.deleted_forms:
                        if item_form.instance.pk:
                            item_form.instance.delete()
                        continue
                    # Only process forms that have changed or are new with data
                    if not item_form.has_changed() and not item_form.instance.pk:
                        continue
                    if item_form.has_changed() or item_form.instance.pk:
                        instance = item_form.save(commit=False)
                        instance.inspection = self.object
                        instance.registered_by = self.request.user
                        instance.save()
                        # item_form.prefix is always the correct Django prefix (e.g. 'items-0')
                        handle_pending_evidences(self.request, instance, prefix=item_form.prefix)
                        # Save M2M if any
                        item_form.save_m2m()
            else:
                transaction.set_rollback(True)
                return self.form_invalid(form)
        return super().form_valid(form)

# 1. Extinguishers
class ExtinguisherListView(LoginRequiredMixin, RolePermissionRequiredMixin, ScheduledInspectionsMixin, ListView):
    permission_required = ('extinguisher', 'view')
    model = ExtinguisherInspection
    template_name = 'inspections/extinguisher_list.html'
    context_object_name = 'inspections'
    inspection_module_type = 'extinguisher'  # For ScheduledInspectionsMixin

    def get_queryset(self):
        qs = super().get_queryset()
        from django.utils import timezone
        current_year = timezone.now().year
        return qs.filter(inspection_date__year=current_year, parent_inspection__isnull=True)
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.utils import timezone
        current_year = timezone.now().year
        
        # Filter scheduled inspections by year (using list comprehension since it's a list now)
        if 'scheduled_inspections' in context:
            context['scheduled_inspections'] = [
                item for item in context['scheduled_inspections'] 
                if item.scheduled_date.year == current_year
            ]
            
        return context

class ExtinguisherCreateView(LoginRequiredMixin, RolePermissionRequiredMixin, InspectionFormUserMixin, FormsetMixin, EvidenceMixin, CreateView):
    permission_required = ('extinguisher', 'create')
    model = ExtinguisherInspection
    form_class = ExtinguisherInspectionForm
    formset_class = ExtinguisherItemFormSet
    template_name = 'inspections/extinguisher_form.html'

    def _count_valid_extinguisher_items(self):
        """Cuenta ítems del formset que tienen datos y NO están marcados para eliminar."""
        total = int(self.request.POST.get('items-TOTAL_FORMS', 0))
        count = 0
        for i in range(total):
            is_deleted = self.request.POST.get(f'items-{i}-DELETE')
            status_val = self.request.POST.get(f'items-{i}-status', '').strip()
            if not is_deleted and status_val:
                count += 1
        return count
    
    def get_initial(self):
        initial = super().get_initial()
        schedule_item_id = self.request.GET.get('schedule_item')
        if schedule_item_id:
            try:
                from .models import InspectionSchedule
                obj = InspectionSchedule.objects.get(pk=schedule_item_id)
                initial['area'] = obj.area
            except InspectionSchedule.DoesNotExist:
                pass
        return initial

    def _sync_extintor_recargas(self):
        """Actualiza ExtintorDetail en inventario si se registró una nueva recarga."""
        from dateutil.relativedelta import relativedelta
        from gestion_activos.models import ExtintorDetail
        actualizados = 0
        for item in self.object.items.filter(fecha_recarga_realizada__isnull=False).select_related('asset__extintor_detail'):
            if not item.asset_id:
                continue
            try:
                detail = item.asset.extintor_detail
                detail.fecha_recarga = item.fecha_recarga_realizada
                if item.fecha_proxima_recarga:
                    detail.fecha_vencimiento = item.fecha_proxima_recarga
                else:
                    detail.fecha_vencimiento = item.fecha_recarga_realizada + relativedelta(years=1)
                detail.save(update_fields=['fecha_recarga', 'fecha_vencimiento'])
                actualizados += 1
            except ExtintorDetail.DoesNotExist:
                pass
        return actualizados

    def form_valid(self, form):
        # --- VALIDACIÓN OBLIGATORIA: Al menos un ítem de detalle ---
        if self._count_valid_extinguisher_items() == 0:
            messages.error(self.request, 'Debe agregar al menos un ítem para guardar la inspección.')
            return self.form_invalid(form)
        # --- FIN VALIDACIÓN ---

        form.instance.inspector = self.request.user

        schedule_item_id = self.request.GET.get('schedule_item')
        if schedule_item_id:
            try:
                from .models import InspectionSchedule
                obj = InspectionSchedule.objects.get(pk=schedule_item_id)
                form.instance.schedule_item = obj
                obj.status = 'Realizada'
                obj.save()
            except InspectionSchedule.DoesNotExist:
                pass

        # Enforce inspector role based on user profile if it matches a valid choice
        user_role_name = self.request.user.get_role_name()
        if user_role_name in [c[0] for c in form.fields['inspector_role'].choices]:
            form.instance.inspector_role = user_role_name

        if form.instance.status == 'Programada':
            form.instance.status = 'En proceso'
        response = super().form_valid(form)

        # Sync recharge dates to inventory
        n = self._sync_extintor_recargas()
        if n > 0:
            messages.info(self.request, f'✅ {n} extintor(es) actualizado(s) en inventario con nueva fecha de recarga.')

        messages.success(self.request, f'Inspección de extintores guardada exitosamente para {form.instance.area}')
        return response
    
    def get_success_url(self):
        return reverse('extinguisher_detail', kwargs={'pk': self.object.pk})

class ExtinguisherUpdateView(LoginRequiredMixin, RolePermissionRequiredMixin, InspectionFormUserMixin, FormsetMixin, EvidenceMixin, UpdateView):
    permission_required = ('extinguisher', 'edit')
    model = ExtinguisherInspection
    form_class = ExtinguisherInspectionForm
    formset_class = ExtinguisherItemFormSet
    template_name = 'inspections/extinguisher_form.html'

    def _count_surviving_extinguisher_items(self):
        """Cuenta ítems existentes que NO están marcados para eliminar, más nuevos con datos."""
        total = int(self.request.POST.get('items-TOTAL_FORMS', 0))
        count = 0
        for i in range(total):
            is_deleted = self.request.POST.get(f'items-{i}-DELETE')
            item_id = self.request.POST.get(f'items-{i}-id', '').strip()
            status_val = self.request.POST.get(f'items-{i}-status', '').strip()
            # Existing item not deleted, OR new item with data
            if not is_deleted and (item_id or status_val):
                count += 1
        return count
    
    def _sync_extintor_recargas(self):
        """Actualiza ExtintorDetail en inventario si se registró una nueva recarga."""
        from dateutil.relativedelta import relativedelta
        from gestion_activos.models import ExtintorDetail
        actualizados = 0
        for item in self.object.items.filter(fecha_recarga_realizada__isnull=False).select_related('asset__extintor_detail'):
            if not item.asset_id:
                continue
            try:
                detail = item.asset.extintor_detail
                detail.fecha_recarga = item.fecha_recarga_realizada
                if item.fecha_proxima_recarga:
                    detail.fecha_vencimiento = item.fecha_proxima_recarga
                else:
                    detail.fecha_vencimiento = item.fecha_recarga_realizada + relativedelta(years=1)
                detail.save(update_fields=['fecha_recarga', 'fecha_vencimiento'])
                actualizados += 1
            except ExtintorDetail.DoesNotExist:
                pass
        return actualizados

    def form_valid(self, form):
        # --- VALIDACIÓN OBLIGATORIA: Al menos un ítem de detalle ---
        if self._count_surviving_extinguisher_items() == 0:
            messages.error(self.request, 'Debe agregar al menos un ítem para guardar la inspección.')
            return self.form_invalid(form)
        # --- FIN VALIDACIÓN ---

        # Enforce inspector role on edit too
        user_role_name = self.request.user.get_role_name()
        if user_role_name in [c[0] for c in form.fields['inspector_role'].choices]:
            form.instance.inspector_role = user_role_name

        if form.instance.status == 'Programada':
            form.instance.status = 'En proceso'
        response = super().form_valid(form)

        # Sync recharge dates to inventory
        n = self._sync_extintor_recargas()
        if n > 0:
            messages.info(self.request, f'✅ {n} extintor(es) actualizado(s) en inventario con nueva fecha de recarga.')

        messages.success(self.request, 'Inspección de extintores actualizada correctamente')
        return response
    
    def get_success_url(self):
        return reverse('extinguisher_detail', kwargs={'pk': self.object.pk})

class ExtinguisherDetailView(LoginRequiredMixin, RolePermissionRequiredMixin, EvidenceMixin, DetailView):
    permission_required = ('extinguisher', 'view')
    model = ExtinguisherInspection
    template_name = 'inspections/extinguisher_detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inspection = self.object
        user = self.request.user
        
        # Timeline Logic
        root = inspection
        while root.parent_inspection:
            root = root.parent_inspection
        timeline = [root]
        def get_descendants(parent):
            children = parent.follow_ups.all().order_by('inspection_date', 'pk')
            for child in children:
                timeline.append(child)
                get_descendants(child)
        get_descendants(root)
        if len(timeline) > 1:
            context['timeline_inspections'] = timeline
            
        # Robust Participants & Signatures Logic
        participants = inspection.get_participants()
        signatures = inspection.signatures.select_related('user').all()
        signatures_dict = {s.user_id: s for s in signatures}
        
        participants_data = []
        for p in participants:
            sig = signatures_dict.get(p.id)
            participants_data.append({
                'user': p,
                'has_signed': sig is not None,
                'signed_at': sig.signed_at if sig else None,
            })
            
        context['participants_data'] = participants_data
        context['signatures'] = signatures
        context['user_has_signed'] = user.id in signatures_dict
        context['is_participant'] = user in participants
        context['has_signature_profile'] = bool(getattr(user, 'digital_signature', None))
        context['can_sign'] = context['is_participant'] and not context['user_has_signed'] and context['has_signature_profile']
        
        return context

class SignExtinguisherInspectionView(LoginRequiredMixin, View):
    def post(self, request, pk):
        inspection = get_object_or_404(ExtinguisherInspection, pk=pk)
        user = request.user
        
        # STEP 1: Participant Validation
        participants = inspection.get_participants()
        if user not in participants:
             messages.error(request, 'No está en la lista de participantes.')
             return redirect('extinguisher_detail', pk=pk)
             
        # DUPLICATE VALIDATION
        if inspection.signatures.filter(user=user).exists():
             messages.warning(request, 'Firmado (ya ha realizado su firma previamente).')
             return redirect('extinguisher_detail', pk=pk)

        # STEP 2: Signature Profile Validation
        if not getattr(user, 'digital_signature', None):
             messages.error(request, 'Debe registrar su firma en el perfil.')
             return redirect('extinguisher_detail', pk=pk)
             
        if inspection.status in ['Cerrada', 'Cerrada con Hallazgos']:
             messages.error(request, 'La inspección ya está cerrada.')
             return redirect('extinguisher_detail', pk=pk)

        # Record signature
        InspectionSignature.objects.create(
            inspection=inspection,
            user=user,
            signature=user.digital_signature
        )
        
        # CHECK IF ALL SIGNED
        all_participants = inspection.get_participants()
        signatures_count = inspection.signatures.count()
        
        if signatures_count >= all_participants.count():
            # FULL CLOSURE LOGIC
            failed_items = inspection.items.filter(status__in=['Malo', 'Recargar'])
            
            if failed_items.exists():
                inspection.status = 'Seguimiento en proceso'
                inspection.save()
                
                # Generación de Seguimiento
                from system_config.models import SystemConfig
                days = SystemConfig.get_value('dias_seguimiento_auto', 15)
                follow_up_date = timezone.now().date() + timedelta(days=days)
                
                follow_up = ExtinguisherInspection.objects.create(
                    parent_inspection=inspection,
                    area=inspection.area,
                    asset=inspection.asset,
                    inspector=inspection.inspector, 
                    inspector_role=inspection.inspector_role,
                    inspection_date=follow_up_date,
                    status='Programada',
                )
                
                for item in failed_items:
                    target_asset_id = item.asset_id or inspection.asset_id
                    new_item = ExtinguisherItem.objects.create(
                        inspection=follow_up,
                        asset_id=target_asset_id,
                        pressure_gauge_ok=item.pressure_gauge_ok,
                        safety_pin_ok=item.safety_pin_ok,
                        hose_nozzle_ok=item.hose_nozzle_ok,
                        signage_ok=item.signage_ok,
                        access_ok=item.access_ok,
                        label_ok=item.label_ok,
                        status='Malo',
                        observations=f"Seguimiento: {item.observations}"[:255] if item.observations else "Seguimiento pendiente",
                        registered_by=user
                    )
                    # Copiar evidencias del ítem original al nuevo ítem de seguimiento
                    from django.contrib.contenttypes.models import ContentType
                    import shutil, os
                    from django.core.files.base import ContentFile
                    item_ct = ContentType.objects.get_for_model(ExtinguisherItem)
                    for evidence in InspectionEvidence.objects.filter(content_type=item_ct, object_id=item.pk):
                        try:
                            evidence.image.open('rb')
                            file_content = evidence.image.read()
                            evidence.image.close()
                            orig_name = os.path.basename(evidence.image.name)
                            InspectionEvidence.objects.create(
                                content_type=item_ct,
                                object_id=new_item.pk,
                                description=evidence.description,
                                uploaded_by=evidence.uploaded_by,
                                image=ContentFile(file_content, name=orig_name)
                            )
                        except Exception:
                            pass  # Si la imagen no se puede copiar, continuar sin bloquear
                
                try:
                    jefes_group = NotificationGroup.objects.get(name="Jefes", is_active=True)
                    for boss in jefes_group.users.all():
                        Notification.objects.create(
                            user=boss,
                            title=f"Hallazgos Críticos - Inspección #{inspection.pk}",
                            message=f"Se ha cerrado la inspección de extintores en {inspection.area} con hallazgos. Se generó seguimiento automático.",
                            link=reverse('extinguisher_detail', kwargs={'pk': follow_up.pk}),
                            notification_type='alert'
                        )
                except: pass
                messages.success(request, f'Inspección finalizada con hallazgos. Seguimiento generado para {follow_up_date.strftime("%d/%m/%Y")}.')
            else:
                inspection.status = 'Cerrada'
                inspection.save()
                
                # SI ES UN SEGUIMIENTO, verificar si el padre ya se puede cerrar definitivamente
                if inspection.parent_inspection:
                    curr = inspection
                    parent_updated = False
                    while curr.parent_inspection:
                        p = curr.parent_inspection
                        # Solo cerramos el padre si todos sus hijos directos están en estados terminales
                        if not p.follow_ups.exclude(status__in=['Cerrada', 'Cerrada con seguimientos']).exists():
                            p.status = 'Cerrada con seguimientos'
                            p.save()
                            curr = p
                            parent_updated = True
                        else:
                            break
                    if parent_updated:
                        messages.success(request, 'Seguimiento completado y jerarquía de inspecciones cerrada.')
                    else:
                        messages.success(request, 'Seguimiento completado.')
                else:
                    messages.success(request, 'Inspección cerrada exitosamente.')
        else:
            inspection.status = 'En proceso'
            inspection.save()
            messages.success(request, 'Firma registrada exitosamente. Pendiente de firmas de otros participantes.')
            
        return redirect('extinguisher_detail', pk=pk)

class ExtinguisherReportView(LoginRequiredMixin, EvidenceMixin, DetailView):
    model = ExtinguisherInspection
    template_name = 'inspections/extinguisher_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = self.object.items.all()
        context['signatures'] = self.object.signatures.select_related('user').all()
        context['total_inspected'] = items.count()
        context['total_good'] = items.filter(status='Bueno').count()
        context['total_bad'] = items.filter(status__in=['Malo', 'Recargar']).count()
        context['any_item_has_evidence'] = any(item.evidences.exists() for item in items)
        return context

class ExtinguisherItemCreateView(LoginRequiredMixin, EvidenceMixin, CreateView):
    model = ExtinguisherItem; form_class = ExtinguisherItemForm; template_name = 'inspections/extinguisher_item_form.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['inspection'] = get_object_or_404(ExtinguisherInspection, pk=self.kwargs['pk'])
        return context
    def form_valid(self, form):
        insp = get_object_or_404(ExtinguisherInspection, pk=self.kwargs['pk'])
        form.instance.inspection = insp
        form.instance.registered_by = self.request.user
        form.save()
        handle_pending_evidences(self.request, form.instance)
        if 'save_and_add' in self.request.POST: return redirect('extinguisher_item_create', pk=insp.pk)
        return redirect('extinguisher_detail', pk=insp.pk)

class ExtinguisherItemUpdateView(LoginRequiredMixin, EvidenceMixin, UpdateView):
    model = ExtinguisherItem; form_class = ExtinguisherItemForm; template_name = 'inspections/extinguisher_item_form.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['inspection'] = self.object.inspection
        return context
    def form_valid(self, form):
        form.instance.registered_by = self.request.user
        return super().form_valid(form)
    def get_success_url(self): return reverse('extinguisher_detail', kwargs={'pk': self.object.inspection.pk})

# 2. First Aid
class FirstAidListView(LoginRequiredMixin, RolePermissionRequiredMixin, ScheduledInspectionsMixin, ListView):
    permission_required = ('first_aid', 'view')
    model = FirstAidInspection
    template_name = 'inspections/first_aid_list.html'
    context_object_name = 'inspections'
    inspection_module_type = 'first_aid'

    def get_queryset(self):
        qs = super().get_queryset()
        from django.utils import timezone
        current_year = timezone.now().year
        return qs.filter(inspection_date__year=current_year, parent_inspection__isnull=True)

class FirstAidCreateView(LoginRequiredMixin, RolePermissionRequiredMixin, InspectionFormUserMixin, FormsetMixin, EvidenceMixin, CreateView):
    permission_required = ('first_aid', 'create')
    model = FirstAidInspection
    form_class = FirstAidInspectionForm
    formset_class = FirstAidItemFormSet
    template_name = 'inspections/first_aid_form.html'

    def _count_valid_firstaid_items(self):
        """Cuenta ítems del formset que tienen datos y NO están marcados para eliminar."""
        total = int(self.request.POST.get('items-TOTAL_FORMS', 0))
        count = 0
        for i in range(total):
            is_deleted = self.request.POST.get(f'items-{i}-DELETE')
            name = self.request.POST.get(f'items-{i}-element_name', '').strip()
            if not is_deleted and name:
                count += 1
        return count
    
    def form_valid(self, form):
        # --- VALIDACIÓN OBLIGATORIA: Al menos un ítem de detalle ---
        if self._count_valid_firstaid_items() == 0:
            messages.error(self.request, 'Debe agregar al menos un ítem para guardar la inspección.')
            return self.form_invalid(form)
        # --- FIN VALIDACIÓN ---

        form.instance.inspector = self.request.user
        # Enforce inspector role
        user_role_name = self.request.user.get_role_name()
        if user_role_name in [c[0] for c in form.fields['inspector_role'].choices]:
            form.instance.inspector_role = user_role_name
            
        if form.instance.status == 'Programada':
            form.instance.status = 'En proceso'
        response = super().form_valid(form)
        messages.success(self.request, f'Inspección de botiquines guardada exitosamente para {form.instance.area}')
        return response
    
    def get_success_url(self):
        return reverse('first_aid_detail', kwargs={'pk': self.object.pk})

class FirstAidUpdateView(LoginRequiredMixin, RolePermissionRequiredMixin, InspectionFormUserMixin, FormsetMixin, EvidenceMixin, UpdateView):
    permission_required = ('first_aid', 'edit')
    model = FirstAidInspection
    form_class = FirstAidInspectionForm
    formset_class = FirstAidItemFormSet
    template_name = 'inspections/first_aid_form.html'

    def _count_surviving_firstaid_items(self):
        """Cuenta ítems existentes que NO están marcados para eliminar, más nuevos con datos."""
        total = int(self.request.POST.get('items-TOTAL_FORMS', 0))
        count = 0
        for i in range(total):
            is_deleted = self.request.POST.get(f'items-{i}-DELETE')
            item_id = self.request.POST.get(f'items-{i}-id', '').strip()
            name = self.request.POST.get(f'items-{i}-element_name', '').strip()
            # Existing item not deleted, OR new item with a name
            if not is_deleted and (item_id or name):
                count += 1
        return count
    
    def form_valid(self, form):
        # --- VALIDACIÓN OBLIGATORIA: Al menos un ítem de detalle ---
        if self._count_surviving_firstaid_items() == 0:
            messages.error(self.request, 'Debe agregar al menos un ítem para guardar la inspección.')
            return self.form_invalid(form)
        # --- FIN VALIDACIÓN ---

        # Enforce inspector role
        user_role_name = self.request.user.get_role_name()
        if user_role_name in [c[0] for c in form.fields['inspector_role'].choices]:
            form.instance.inspector_role = user_role_name
            
        if form.instance.status == 'Programada':
            form.instance.status = 'En proceso'
        response = super().form_valid(form)
        messages.success(self.request, 'Inspección de botiquines actualizada correctamente')
        return response
    
    def get_success_url(self):
        return reverse('first_aid_detail', kwargs={'pk': self.object.pk})

class FirstAidDetailView(LoginRequiredMixin, RolePermissionRequiredMixin, EvidenceMixin, DetailView):
    permission_required = ('first_aid', 'view')
    model = FirstAidInspection
    template_name = 'inspections/first_aid_detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inspection = self.object
        user = self.request.user
        
        # Timeline Logic
        root = inspection
        while root.parent_inspection:
            root = root.parent_inspection
            
        timeline = [root]
        
        def get_descendants(parent):
            children = parent.follow_ups.all().order_by('inspection_date', 'pk')
            for child in children:
                timeline.append(child)
                get_descendants(child)
                
        get_descendants(root)
        
        if len(timeline) > 1:
            context['timeline_inspections'] = timeline
            
        # Robust Participants & Signatures Logic
        participants = inspection.get_participants()
        signatures = inspection.signatures.select_related('user').all()
        signatures_dict = {s.user_id: s for s in signatures}
        
        participants_data = []
        for p in participants:
            sig = signatures_dict.get(p.id)
            participants_data.append({
                'user': p,
                'has_signed': sig is not None,
                'signed_at': sig.signed_at if sig else None,
            })
            
        context['participants_data'] = participants_data
        context['signatures'] = signatures
        context['user_has_signed'] = user.id in signatures_dict
        context['is_participant'] = user in participants
        context['has_signature_profile'] = bool(getattr(user, 'digital_signature', None))
        context['can_sign'] = context['is_participant'] and not context['user_has_signed'] and context['has_signature_profile']
        
        return context

class SignFirstAidInspectionView(LoginRequiredMixin, View):
    def post(self, request, pk):
        inspection = get_object_or_404(FirstAidInspection, pk=pk)
        user = request.user
        
        # STEP 1: Participant Validation
        participants = inspection.get_participants()
        if user not in participants:
             messages.error(request, 'No está en la lista de participantes.')
             return redirect('first_aid_detail', pk=pk)
             
        # DUPLICATE VALIDATION
        if inspection.signatures.filter(user=user).exists():
             messages.warning(request, 'Firmado (ya ha realizado su firma previamente).')
             return redirect('first_aid_detail', pk=pk)

        # STEP 2: Signature Profile Validation
        if not getattr(user, 'digital_signature', None):
             messages.error(request, 'Debe registrar su firma en el perfil.')
             return redirect('first_aid_detail', pk=pk)
             
        if inspection.status in ['Cerrada', 'Cerrada con Hallazgos']:
             messages.error(request, 'La inspección ya está cerrada.')
             return redirect('first_aid_detail', pk=pk)

        # Record signature
        from .models import FirstAidSignature, FirstAidItem
        FirstAidSignature.objects.create(
            inspection=inspection,
            user=user,
            signature=user.digital_signature
        )
        
        # CHECK IF ALL SIGNED
        all_participants = inspection.get_participants()
        signatures_count = inspection.signatures.count()
        
        if signatures_count >= all_participants.count():
            missing_items = inspection.items.filter(status='No Existe')
            if missing_items.exists():
                inspection.status = 'Seguimiento en proceso'
                inspection.general_status = 'No Cumple'
                inspection.save()
                
                from system_config.models import SystemConfig
                days = SystemConfig.get_value('dias_seguimiento_auto', 15)
                follow_up_date = timezone.now().date() + timedelta(days=days)
                follow_up = FirstAidInspection.objects.create(
                    parent_inspection=inspection,
                    area=inspection.area,
                    inspector=inspection.inspector, 
                    inspector_role=inspection.inspector_role,
                    inspection_date=follow_up_date,
                    status='Programada',
                )
                for item in missing_items:
                    new_item = FirstAidItem.objects.create(
                        inspection=follow_up,
                        element_name=item.element_name,
                        quantity=item.quantity,
                        expiration_date=item.expiration_date,
                        status='No Existe',
                        observations=f"Seguimiento: {item.observations}"[:255] if item.observations else "Seguimiento"
                    )
                    # Copy evidences from the parent item to the new follow-up item
                    for evidence in item.evidences.all():
                        InspectionEvidence.objects.create(
                            content_type=evidence.content_type,
                            object_id=new_item.pk,
                            image=evidence.image,
                            description=evidence.description,
                            uploaded_by=evidence.uploaded_by,
                        )
                messages.info(request, f"Inspección finalizada con hallazgos. Seguimiento generado para {follow_up_date.strftime('%d/%m/%Y')}")
            else:
                inspection.status = 'Cerrada'
                inspection.general_status = 'Cumple'
                inspection.save()
                
                # SI ES UN SEGUIMIENTO, verificar si el padre ya se puede cerrar definitivamente
                if inspection.parent_inspection:
                    curr = inspection
                    parent_updated = False
                    while curr.parent_inspection:
                        p = curr.parent_inspection
                        if not p.follow_ups.exclude(status__in=['Cerrada', 'Cerrada con seguimientos']).exists():
                            p.status = 'Cerrada con seguimientos'
                            p.save()
                            curr = p
                            parent_updated = True
                        else:
                            break
                    if parent_updated:
                        messages.success(request, 'Seguimiento completado y jerarquía de inspecciones cerrada.')
                    else:
                        messages.success(request, 'Seguimiento completado.')
                else:
                    if inspection.schedule_item:
                        s_item = inspection.schedule_item
                        s_item.status = 'Realizada'
                        s_item.save()
                        next_s = s_item.generate_next_schedule()
                        if next_s:
                            messages.info(request, f"📅 Nueva programación generada para {next_s.scheduled_date}")
                    messages.success(request, 'Inspección cerrada por completar todas las firmas.')
        else:
            inspection.status = 'En proceso'
            inspection.save()
            messages.success(request, 'Firma registrada. Pendiente de firmas de otros participantes.')
        
        return redirect('first_aid_detail', pk=pk)

class FirstAidReportView(LoginRequiredMixin, EvidenceMixin, DetailView):
    model = FirstAidInspection
    template_name = 'inspections/first_aid_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = self.object.items.all()
        context['items_exist_count'] = items.filter(status='Existe').count()
        context['items_missing_count'] = items.filter(status='No Existe').count()
        context['any_item_has_evidence'] = any(item.evidences.exists() for item in items)
        return context

class FirstAidItemCreateView(LoginRequiredMixin, EvidenceMixin, CreateView):
    model = FirstAidItem; form_class = FirstAidItemForm; template_name = 'inspections/first_aid_item_form.html'
    def form_valid(self, form):
        insp = get_object_or_404(FirstAidInspection, pk=self.kwargs['pk'])
        form.instance.inspection = insp
        form.instance.registered_by = self.request.user
        form.save()
        handle_pending_evidences(self.request, form.instance)
        if 'save_and_add' in self.request.POST: return redirect('first_aid_item_create', pk=insp.pk)
        return redirect('first_aid_detail', pk=insp.pk)

class FirstAidItemUpdateView(LoginRequiredMixin, EvidenceMixin, UpdateView):
    model = FirstAidItem; form_class = FirstAidItemForm; template_name = 'inspections/first_aid_item_form.html'
    def form_valid(self, form):
        form.instance.registered_by = self.request.user
        return super().form_valid(form)
    def get_success_url(self): return reverse('first_aid_detail', kwargs={'pk': self.object.inspection.pk})

# 3. Process
class ProcessListView(LoginRequiredMixin, RolePermissionRequiredMixin, ScheduledInspectionsMixin, ListView):
    permission_required = ('process', 'view')
    model = ProcessInspection
    template_name = 'inspections/process_list.html'
    context_object_name = 'inspections'
    inspection_module_type = 'process'

    def get_queryset(self):
        qs = super().get_queryset()
        from django.utils import timezone
        current_year = timezone.now().year
        return qs.filter(inspection_date__year=current_year, parent_inspection__isnull=True)
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.utils import timezone
        current_year = timezone.now().year
        
        if 'scheduled_inspections' in context:
            context['scheduled_inspections'] = [
                item for item in context['scheduled_inspections'] 
                if item.scheduled_date.year == current_year
            ]
            
        return context

class ProcessCreateView(LoginRequiredMixin, RolePermissionRequiredMixin, InspectionFormUserMixin, FormsetMixin, EvidenceMixin, CreateView):
    permission_required = ('process', 'create')
    model = ProcessInspection
    form_class = ProcessInspectionForm
    formset_class = ProcessItemFormSet
    template_name = 'inspections/process_form.html'
    
    initial_items = [
        "1. ¿Las áreas están señalizadas y la señalización se encuentra en buen estado?",
        "2. ¿Las paredes están limpias y el estado de la pintura es óptimo?",
        "3. ¿Los pisos están sin grietas y en buen estado de limpieza?",
        "4. ¿Existen equipos de control de incendios ubicados en lugar de fácil acceso?",
        "5. ¿Las áreas se encuentran en adecuado orden y aseo?",
        "6. ¿Las áreas cuentan con buena iluminación?",
        "7. ¿Las instalaciones eléctricas no presentan riesgos y tableros en buen estado?",
        "8. ¿Los recipientes de basura o residuos están señalizados?",
        "9. ¿Las áreas operativas están libres de materiales innecesarios?",
        "10. ¿Los lugares de acceso se encuentran libres de obstáculos?",
        "11. El personal que manipula las sustancias químicas conocen las hojas de seguridad (MSDS)?",
        "12. ¿Equipos de seguridad (extintores, camillas, botiquines) señalizados?",
        "13. ¿Recipientes de sustancias químicas rotulados e identificados?",
        "14. ¿Las máquinas y/o equipos del área se encuentran señalizadas?",
        "15. ¿Los sistemas de seguridad de la máquina funcionan y están señalizados?",
        "16. ¿La señalización de evacuación o de emergencia se encuentra en buen estado?"
    ]

    def get_initial(self):
        initial = super().get_initial()
        schedule_item_id = self.request.GET.get('schedule_item')
        if schedule_item_id:
            try:
                from .models import InspectionSchedule
                obj = InspectionSchedule.objects.get(pk=schedule_item_id)
                initial['area'] = obj.area
            except InspectionSchedule.DoesNotExist:
                pass
        return initial

    def form_valid(self, form):
        form.instance.inspector = self.request.user
        
        schedule_item_id = self.request.GET.get('schedule_item')
        if schedule_item_id:
            try:
                from .models import InspectionSchedule
                obj = InspectionSchedule.objects.get(pk=schedule_item_id)
                form.instance.schedule_item = obj
                obj.status = 'Realizada'
                obj.save()
            except InspectionSchedule.DoesNotExist:
                pass

        # Enforce inspector role
        user_role_name = self.request.user.get_role_name()
        if user_role_name in [c[0] for c in form.fields['inspector_role'].choices]:
            form.instance.inspector_role = user_role_name
            
        if form.instance.status == 'Programada':
            form.instance.status = 'En proceso'
        response = super().form_valid(form)
        messages.success(self.request, f'Inspección de procesos guardada exitosamente para {form.instance.area}')
        return response

    def get_success_url(self):
        return reverse('process_detail', kwargs={'pk': self.object.pk})

class ProcessUpdateView(LoginRequiredMixin, RolePermissionRequiredMixin, InspectionFormUserMixin, FormsetMixin, EvidenceMixin, UpdateView):
    permission_required = ('process', 'edit')
    model = ProcessInspection
    form_class = ProcessInspectionForm
    formset_class = ProcessItemFormSet
    template_name = 'inspections/process_form.html'
    
    def form_valid(self, form):
        # Enforce inspector role
        user_role_name = self.request.user.get_role_name()
        if user_role_name in [c[0] for c in form.fields['inspector_role'].choices]:
            form.instance.inspector_role = user_role_name
            
        if form.instance.status == 'Programada':
            form.instance.status = 'En proceso'
        response = super().form_valid(form)
        messages.success(self.request, 'Inspección de procesos actualizada correctamente')
        return response
    
    def get_success_url(self):
        return reverse('process_detail', kwargs={'pk': self.object.pk})

class ProcessItemUpdateView(LoginRequiredMixin, RolePermissionRequiredMixin, EvidenceMixin, UpdateView):
    permission_required = ('process', 'edit')
    model = ProcessCheckItem
    form_class = ProcessCheckItemForm
    template_name = 'inspections/process_item_form.html'
    
    def form_valid(self, form):
        form.instance.registered_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('process_detail', kwargs={'pk': self.object.inspection.pk})

class ProcessDetailView(LoginRequiredMixin, RolePermissionRequiredMixin, EvidenceMixin, DetailView):
    permission_required = ('process', 'view')
    model = ProcessInspection
    template_name = 'inspections/process_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inspection = self.object
        user = self.request.user
        
        # Timeline Logic
        root = inspection
        while root.parent_inspection:
            root = root.parent_inspection
        timeline = [root]
        def get_descendants(parent):
            children = parent.follow_ups.all().order_by('inspection_date', 'pk')
            for child in children:
                timeline.append(child)
                get_descendants(child)
        get_descendants(root)
        if len(timeline) > 1:
            context['timeline_inspections'] = timeline
            
        # Robust Participants & Signatures Logic
        participants = inspection.get_participants()
        signatures = inspection.signatures.select_related('user').all()
        signatures_dict = {s.user_id: s for s in signatures}
        
        participants_data = []
        for p in participants:
            sig = signatures_dict.get(p.id)
            participants_data.append({
                'user': p,
                'has_signed': sig is not None,
                'signed_at': sig.signed_at if sig else None,
            })
            
        context['participants_data'] = participants_data
        context['signatures'] = signatures
        context['user_has_signed'] = user.id in signatures_dict
        context['is_participant'] = user in participants
        context['has_signature_profile'] = bool(getattr(user, 'digital_signature', None))
        context['can_sign'] = context['is_participant'] and not context['user_has_signed'] and context['has_signature_profile']
        
        return context

class SignProcessInspectionView(LoginRequiredMixin, View):
    def post(self, request, pk):
        inspection = get_object_or_404(ProcessInspection, pk=pk)
        user = request.user
        
        # STEP 1: Participant Validation
        participants = inspection.get_participants()
        if user not in participants:
             messages.error(request, 'No está en la lista de participantes.')
             return redirect('process_detail', pk=pk)
             
        # DUPLICATE VALIDATION
        if inspection.signatures.filter(user=user).exists():
             messages.warning(request, 'Firmado (ya ha realizado su firma previamente).')
             return redirect('process_detail', pk=pk)

        # STEP 2: Signature Profile Validation
        if not getattr(user, 'digital_signature', None):
             messages.error(request, 'Debe registrar su firma en el perfil.')
             return redirect('process_detail', pk=pk)
             
        if inspection.status in ['Cerrada', 'Cerrada con Hallazgos']:
             messages.error(request, 'La inspección ya está cerrada.')
             return redirect('process_detail', pk=pk)

        # Record signature
        ProcessSignature.objects.create(
            inspection=inspection,
            user=user,
            signature=user.digital_signature
        )
        
        # CHECK IF ALL SIGNED
        all_participants = inspection.get_participants()
        signatures_count = inspection.signatures.count()
        
        if signatures_count >= all_participants.count():
            failed_items = inspection.items.filter(item_status='Malo')
            if failed_items.exists():
                inspection.status = 'Seguimiento en proceso'
                inspection.save()
                
                from system_config.models import SystemConfig
                days = SystemConfig.get_value('dias_seguimiento_auto', 15)
                follow_up_date = timezone.now().date() + timedelta(days=days)
                
                follow_up = ProcessInspection.objects.create(
                    parent_inspection=inspection,
                    area=inspection.area,
                    inspector=inspection.inspector, 
                    inspector_role=inspection.inspector_role,
                    inspected_process=inspection.inspected_process,
                    inspection_date=follow_up_date,
                    status='Programada',
                )
                for item in failed_items:
                    ProcessCheckItem.objects.create(
                        inspection=follow_up,
                        question=item.question,
                        response='No',
                        item_status='Malo',
                        observations=f"Seguimiento: {item.observations}"[:255] if item.observations else "Seguimiento"
                    )
                messages.info(request, f"Inspección finalizada con hallazgos. Seguimiento generado para {follow_up_date.strftime('%d/%m/%Y')}")
            else:
                inspection.status = 'Cerrada'
                inspection.save()
                
                # SI ES UN SEGUIMIENTO, verificar si el padre ya se puede cerrar definitivamente
                if inspection.parent_inspection:
                    curr = inspection
                    parent_updated = False
                    while curr.parent_inspection:
                        p = curr.parent_inspection
                        if not p.follow_ups.exclude(status__in=['Cerrada', 'Cerrada con seguimientos']).exists():
                            p.status = 'Cerrada con seguimientos'
                            p.save()
                            curr = p
                            parent_updated = True
                        else:
                            break
                    if parent_updated:
                        messages.success(request, 'Seguimiento completado y jerarquía de inspecciones cerrada.')
                    else:
                        messages.success(request, 'Seguimiento completado.')
                else:
                    if inspection.schedule_item:
                        s_item = inspection.schedule_item
                        s_item.status = 'Realizada'
                        s_item.save()
                        next_s = s_item.generate_next_schedule()
                        if next_s:
                            messages.info(request, f"📅 Nueva programación generada para {next_s.scheduled_date}")
                    messages.success(request, 'Inspección cerrada exitosamente.')
        else:
            inspection.status = 'En proceso'
            inspection.save()
            messages.success(request, 'Firma registrada. Pendiente de firmas de otros participantes.')
        
        return redirect('process_detail', pk=pk)

class ProcessReportView(LoginRequiredMixin, EvidenceMixin, DetailView):
    model = ProcessInspection
    template_name = 'inspections/process_report.html'

# 4. Storage
class StorageListView(LoginRequiredMixin, RolePermissionRequiredMixin, ScheduledInspectionsMixin, ListView):
    permission_required = ('storage', 'view')
    model = StorageInspection
    template_name = 'inspections/storage_list.html'
    context_object_name = 'inspections'
    inspection_module_type = 'storage'

    def get_queryset(self):
        qs = super().get_queryset()
        from django.utils import timezone
        current_year = timezone.now().year
        return qs.filter(inspection_date__year=current_year, parent_inspection__isnull=True)
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.utils import timezone
        current_year = timezone.now().year
        
        if 'scheduled_inspections' in context:
            context['scheduled_inspections'] = [
                item for item in context['scheduled_inspections'] 
                if item.scheduled_date.year == current_year
            ]
            
        return context

class StorageCreateView(LoginRequiredMixin, RolePermissionRequiredMixin, InspectionFormUserMixin, FormsetMixin, EvidenceMixin, CreateView):
    permission_required = ('storage', 'create')
    model = StorageInspection
    form_class = StorageInspectionForm
    formset_class = StorageItemFormSet
    template_name = 'inspections/storage_form.html'
    
    initial_items = [
        "1. ¿Las áreas están claramente señalizadas y la señalización se encuentra en buen estado?",
        "2. ¿Las paredes están limpias y el estado de la pintura es óptimo?",
        "3. ¿Los pisos están sin grietas y en buen estado de limpieza?",
        "4. ¿Existen equipos de control de incendios ubicados en lugar de fácil acceso?",
        "5. ¿Las áreas se encuentran en adecuado orden y aseo?",
        "6. ¿Las áreas cuentan con buena iluminación?",
        "7. ¿Las instalaciones eléctricas no presentan riesgos y tableros en buen estado?",
        "8. ¿Los recipientes de basura o residuos están señalizados?",
        "9. ¿La distribución de estantes permite la circulación por los pasillos?",
        "10. ¿Los artículos de mayor peso se almacenan en la parte inferior?",
        "11. ¿El personal conoce las hojas de seguridad (MSDS)?",
        "12. ¿Equipos de seguridad (extintores, camillas, botiquines) señalizados?",
        "13. ¿Recipientes de sustancias químicas rotulados e identificados?",
        "14. ¿Equipos de transporte (estibadores) en buenas condiciones?",
        "15. ¿La señalización de evacuación o emergencia en buen estado?"
    ]

    def get_initial(self):
        initial = super().get_initial()
        schedule_item_id = self.request.GET.get('schedule_item')
        if schedule_item_id:
            try:
                from .models import InspectionSchedule
                obj = InspectionSchedule.objects.get(pk=schedule_item_id)
                initial['area'] = obj.area
            except InspectionSchedule.DoesNotExist:
                pass
        return initial

    def form_valid(self, form):
        form.instance.inspector = self.request.user
        
        schedule_item_id = self.request.GET.get('schedule_item')
        if schedule_item_id:
            try:
                from .models import InspectionSchedule
                obj = InspectionSchedule.objects.get(pk=schedule_item_id)
                form.instance.schedule_item = obj
                obj.status = 'Realizada'
                obj.save()
            except InspectionSchedule.DoesNotExist:
                pass

        # Enforce inspector role
        user_role_name = self.request.user.get_role_name()
        if user_role_name in [c[0] for c in form.fields['inspector_role'].choices]:
            form.instance.inspector_role = user_role_name
            
        if form.instance.status == 'Programada':
            form.instance.status = 'En proceso'
        response = super().form_valid(form)
        messages.success(self.request, f'Inspección de almacenamiento guardada exitosamente para {form.instance.area}')
        return response

    def get_success_url(self):
        return reverse('storage_detail', kwargs={'pk': self.object.pk})

class StorageUpdateView(LoginRequiredMixin, RolePermissionRequiredMixin, InspectionFormUserMixin, FormsetMixin, EvidenceMixin, UpdateView):
    permission_required = ('storage', 'edit')
    model = StorageInspection
    form_class = StorageInspectionForm
    formset_class = StorageItemFormSet
    template_name = 'inspections/storage_form.html'
    
    def form_valid(self, form):
        # Enforce inspector role
        user_role_name = self.request.user.get_role_name()
        if user_role_name in [c[0] for c in form.fields['inspector_role'].choices]:
            form.instance.inspector_role = user_role_name
            
        if form.instance.status == 'Programada':
            form.instance.status = 'En proceso'
        response = super().form_valid(form)
        messages.success(self.request, 'Inspección de almacenamiento actualizada correctamente')
        return response
    
    def get_success_url(self):
        return reverse('storage_detail', kwargs={'pk': self.object.pk})

class StorageItemUpdateView(LoginRequiredMixin, RolePermissionRequiredMixin, EvidenceMixin, UpdateView):
    permission_required = ('storage', 'edit')
    model = StorageCheckItem
    form_class = StorageCheckItemForm
    template_name = 'inspections/storage_item_form.html'
    
    def form_valid(self, form):
        form.instance.registered_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('storage_detail', kwargs={'pk': self.object.inspection.pk})

class StorageDetailView(LoginRequiredMixin, RolePermissionRequiredMixin, EvidenceMixin, DetailView):
    permission_required = ('storage', 'view')
    model = StorageInspection
    template_name = 'inspections/storage_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inspection = self.object
        user = self.request.user
        
        root = inspection
        while root.parent_inspection:
            root = root.parent_inspection
        
        timeline = [root]
        def get_descendants(parent):
            children = parent.follow_ups.all().order_by('inspection_date', 'pk')
            for child in children:
                timeline.append(child)
                get_descendants(child)
        get_descendants(root)
        if len(timeline) > 1:
            context['timeline_inspections'] = timeline
            
        # Robust Participants & Signatures Logic
        participants = inspection.get_participants()
        signatures = inspection.signatures.select_related('user').all()
        signatures_dict = {s.user_id: s for s in signatures}
        
        participants_data = []
        for p in participants:
            sig = signatures_dict.get(p.id)
            participants_data.append({
                'user': p,
                'has_signed': sig is not None,
                'signed_at': sig.signed_at if sig else None,
            })
            
        context['participants_data'] = participants_data
        context['signatures'] = signatures
        context['user_has_signed'] = user.id in signatures_dict
        context['is_participant'] = user in participants
        context['has_signature_profile'] = bool(getattr(user, 'digital_signature', None))
        context['can_sign'] = context['is_participant'] and not context['user_has_signed'] and context['has_signature_profile']
        
        return context

class SignStorageInspectionView(LoginRequiredMixin, View):
    def post(self, request, pk):
        inspection = get_object_or_404(StorageInspection, pk=pk)
        user = request.user
        
        # STEP 1: Participant Validation
        participants = inspection.get_participants()
        if user not in participants:
             messages.error(request, 'No está en la lista de participantes.')
             return redirect('storage_detail', pk=pk)
             
        # DUPLICATE VALIDATION
        if inspection.signatures.filter(user=user).exists():
             messages.warning(request, 'Firmado (ya ha realizado su firma previamente).')
             return redirect('storage_detail', pk=pk)

        # STEP 2: Signature Profile Validation
        if not getattr(user, 'digital_signature', None):
             messages.error(request, 'Debe registrar su firma en el perfil.')
             return redirect('storage_detail', pk=pk)
             
        if inspection.status in ['Cerrada', 'Cerrada con Hallazgos']:
             messages.error(request, 'La inspección ya está cerrada.')
             return redirect('storage_detail', pk=pk)

        # Record signature
        StorageSignature.objects.create(
            inspection=inspection,
            user=user,
            signature=user.digital_signature
        )
        
        # CHECK IF ALL SIGNED
        all_participants = inspection.get_participants()
        signatures_count = inspection.signatures.count()
        
        if signatures_count >= all_participants.count():
            failed_items = inspection.items.filter(item_status='Malo')
            if failed_items.exists():
                inspection.status = 'Seguimiento en proceso'
                inspection.save()
                
                from system_config.models import SystemConfig
                days = SystemConfig.get_value('dias_seguimiento_auto', 15)
                follow_up_date = timezone.now().date() + timedelta(days=days)
                
                follow_up = StorageInspection.objects.create(
                    parent_inspection=inspection,
                    area=inspection.area,
                    inspector=inspection.inspector, 
                    inspector_role=inspection.inspector_role,
                    inspected_process=inspection.inspected_process,
                    inspection_date=follow_up_date,
                    status='Programada',
                )
                for item in failed_items:
                    StorageCheckItem.objects.create(
                        inspection=follow_up,
                        question=item.question,
                        response='No',
                        item_status='Malo',
                        observations=f"Seguimiento: {item.observations}"[:255] if item.observations else "Seguimiento"
                    )
                messages.info(request, f"Inspección finalizada con hallazgos. Seguimiento generado para {follow_up_date.strftime('%d/%m/%Y')}")
            else:
                inspection.status = 'Cerrada'
                inspection.save()
                
                # SI ES UN SEGUIMIENTO, verificar si el padre ya se puede cerrar definitivamente
                if inspection.parent_inspection:
                    curr = inspection
                    parent_updated = False
                    while curr.parent_inspection:
                        p = curr.parent_inspection
                        if not p.follow_ups.exclude(status__in=['Cerrada', 'Cerrada con seguimientos']).exists():
                            p.status = 'Cerrada con seguimientos'
                            p.save()
                            curr = p
                            parent_updated = True
                        else:
                            break
                    if parent_updated:
                        messages.success(request, 'Seguimiento completado y jerarquía de inspecciones cerrada.')
                    else:
                        messages.success(request, 'Seguimiento completado.')
                else:
                    if inspection.schedule_item:
                        s_item = inspection.schedule_item
                        s_item.status = 'Realizada'
                        s_item.save()
                        next_s = s_item.generate_next_schedule()
                        if next_s:
                            messages.info(request, f"📅 Nueva programación generada para {next_s.scheduled_date}")
                    messages.success(request, 'Inspección cerrada exitosamente.')
        else:
            inspection.status = 'En proceso'
            inspection.save()
            messages.success(request, 'Firma registrada. Pendiente de firmas de otros participantes.')
        
        return redirect('storage_detail', pk=pk)

class StorageReportView(LoginRequiredMixin, EvidenceMixin, DetailView):
    model = StorageInspection
    template_name = 'inspections/storage_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['signatures'] = self.object.signatures.select_related('user').all()
        return context

# 5. Forklift
class ForkliftListView(LoginRequiredMixin, RolePermissionRequiredMixin, ScheduledInspectionsMixin, ListView):
    permission_required = ('forklift', 'view')
    model = ForkliftInspection
    template_name = 'inspections/forklift_list.html'
    context_object_name = 'inspections'
    inspection_module_type = 'forklift'
    
    def get_queryset(self):
        qs = super().get_queryset()
        from django.utils import timezone
        current_year = timezone.now().year
        return qs.filter(inspection_date__year=current_year, parent_inspection__isnull=True)

class ForkliftCreateView(LoginRequiredMixin, RolePermissionRequiredMixin, InspectionFormUserMixin, FormsetMixin, EvidenceMixin, CreateView):
    permission_required = ('forklift', 'create')
    model = ForkliftInspection
    form_class = ForkliftInspectionForm
    formset_class = ForkliftItemFormSet
    template_name = 'inspections/forklift_form.html'
    
    initial_items = [
        "1. Bateria sin Novedad (No sulfatada o cargada)",
        "2. Tanque de Gas (Golpes, Fugas)",
        "3. Agarradera trasera (Puntos de apoyo)",
        "4. Tres puntos de apoyo",
        "5. Escalón Antiderrapante",
        "6. Abulladuras, golpes, rayones, varios, etc...",
        "7. Control de velocidad velocidad regulada",
        "8. Acrilico protector del diplay en buen estado",
        "9. Limpieza del equipo",
        "10. Extintor (Buen estado, cargado, vigente)",
        "11. Espejos completos y en buen estado",
        "12. Llantas en buen estado (Delanteras, traseras)",
        "13. Pito Funcionando",
        "14. Asiento funcional (Respaldo y guias)",
        "15. Cinturon de seguridad en buen estado",
        "16. Etiqueta de altura maxima",
        "17. Etiqueta de peso máximo en buen estado",
        "18. Etiqueta de velocidad máxima",
        "19. Rejilla protectora superior en buen estado",
        "23. Luces funcionando (Frontales, traseras, reversa, torreta)",
        "25. Medidor nivel de combustible funcionando",
        "24. Alarma de reversa Funcionando",
        "25. Frenos en buen estado",
        "26. Dirección en buen estado (Sin juego)",
        "27. Funcionamiento del mastil adecuado",
        "28. Freno de mano en buen estado y funcionando",
        "29. Cadena de mastil en buen estado",
        "30. Presenta alguna fuga de aceite",
        "31. ¿El equipo es apto para su funcionamiento?"
    ]

    def form_valid(self, form):
        form.instance.inspector = self.request.user
        
        schedule_item_id = self.request.GET.get('schedule_item')
        if schedule_item_id:
            try:
                from .models import InspectionSchedule
                obj = InspectionSchedule.objects.get(pk=schedule_item_id)
                form.instance.schedule_item = obj
                obj.status = 'Realizada'
                obj.save()
            except InspectionSchedule.DoesNotExist:
                pass

        # Enforce inspector role
        user_role_name = self.request.user.get_role_name()
        if user_role_name in [c[0] for c in form.fields['inspector_role'].choices]:
            form.instance.inspector_role = user_role_name
            
        if form.instance.status == 'Programada':
            form.instance.status = 'En proceso'
        response = super().form_valid(form)
        messages.success(self.request, f'Inspección de montacargas guardada exitosamente - {form.instance.forklift_type}')
        return response

    def get_success_url(self):
        return reverse('forklift_detail', kwargs={'pk': self.object.pk})

class ForkliftUpdateView(LoginRequiredMixin, RolePermissionRequiredMixin, InspectionFormUserMixin, FormsetMixin, EvidenceMixin, UpdateView):
    permission_required = ('forklift', 'edit')
    model = ForkliftInspection
    form_class = ForkliftInspectionForm
    formset_class = ForkliftItemFormSet
    template_name = 'inspections/forklift_form.html'
    
    def form_valid(self, form):
        # Enforce inspector role
        user_role_name = self.request.user.get_role_name()
        if user_role_name in [c[0] for c in form.fields['inspector_role'].choices]:
            form.instance.inspector_role = user_role_name
            
        if form.instance.status == 'Programada':
            form.instance.status = 'En proceso'
        response = super().form_valid(form)
        messages.success(self.request, 'Inspección de montacargas actualizada correctamente')
        return response
    
    def get_success_url(self):
        return reverse('forklift_detail', kwargs={'pk': self.object.pk})

class ForkliftItemUpdateView(LoginRequiredMixin, RolePermissionRequiredMixin, EvidenceMixin, UpdateView):
    permission_required = ('forklift', 'edit')
    model = ForkliftCheckItem
    form_class = ForkliftCheckItemForm
    template_name = 'inspections/forklift_item_form.html'
    
    def form_valid(self, form):
        form.instance.registered_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('forklift_detail', kwargs={'pk': self.object.inspection.pk})

class ForkliftDetailView(LoginRequiredMixin, RolePermissionRequiredMixin, EvidenceMixin, DetailView):
    permission_required = ('forklift', 'view')
    model = ForkliftInspection
    template_name = 'inspections/forklift_detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inspection = self.object
        user = self.request.user
        
        root = inspection
        while root.parent_inspection:
            root = root.parent_inspection
        
        timeline = [root]
        def get_descendants(parent):
            children = parent.follow_ups.all().order_by('inspection_date', 'pk')
            for child in children:
                timeline.append(child)
                get_descendants(child)
        get_descendants(root)
        if len(timeline) > 1:
            context['timeline_inspections'] = timeline
            
        # Robust Participants & Signatures Logic
        participants = inspection.get_participants()
        signatures = inspection.signatures.select_related('user').all()
        signatures_dict = {s.user_id: s for s in signatures}
        
        participants_data = []
        for p in participants:
            sig = signatures_dict.get(p.id)
            participants_data.append({
                'user': p,
                'has_signed': sig is not None,
                'signed_at': sig.signed_at if sig else None,
            })
            
        context['participants_data'] = participants_data
        context['signatures'] = signatures
        context['user_has_signed'] = user.id in signatures_dict
        context['is_participant'] = user in participants
        context['has_signature_profile'] = bool(getattr(user, 'digital_signature', None))
        context['can_sign'] = context['is_participant'] and not context['user_has_signed'] and context['has_signature_profile']
        
        return context

class SignForkliftInspectionView(LoginRequiredMixin, View):
    def post(self, request, pk):
        inspection = get_object_or_404(ForkliftInspection, pk=pk)
        user = request.user
        
        # STEP 1: Participant Validation
        participants = inspection.get_participants()
        if user not in participants:
             messages.error(request, 'No está en la lista de participantes.')
             return redirect('forklift_detail', pk=pk)
             
        # DUPLICATE VALIDATION
        if inspection.signatures.filter(user=user).exists():
             messages.warning(request, 'Firmado (ya ha realizado su firma previamente).')
             return redirect('forklift_detail', pk=pk)

        # STEP 2: Signature Profile Validation
        if not getattr(user, 'digital_signature', None):
             messages.error(request, 'Debe registrar su firma en el perfil.')
             return redirect('forklift_detail', pk=pk)
             
        if inspection.status in ['Cerrada', 'Cerrada con Hallazgos']:
             messages.error(request, 'La inspección ya está cerrada.')
             return redirect('forklift_detail', pk=pk)

        # Record signature
        ForkliftSignature.objects.create(
            inspection=inspection,
            user=user,
            signature=user.digital_signature
        )
        
        # CHECK IF ALL SIGNED
        all_participants = inspection.get_participants()
        signatures_count = inspection.signatures.count()
        
        if signatures_count >= all_participants.count():
            failed_items = inspection.items.filter(item_status='Malo')
            if failed_items.exists():
                inspection.status = 'Seguimiento en proceso'
                inspection.save()
                
                from system_config.models import SystemConfig
                days = SystemConfig.get_value('dias_seguimiento_auto', 15)
                follow_up_date = timezone.now().date() + timedelta(days=days)
                
                follow_up = ForkliftInspection.objects.create(
                    parent_inspection=inspection,
                    area=inspection.area,
                    asset=inspection.asset,
                    inspector=inspection.inspector, 
                    inspector_role=inspection.inspector_role,
                    inspection_date=follow_up_date,
                    status='Programada',
                    forklift_type=inspection.forklift_type
                )
                for item in failed_items:
                    ForkliftCheckItem.objects.create(
                        inspection=follow_up,
                        question=item.question,
                        response=item.response,
                        item_status='Malo',
                        observations=f"Seguimiento: {item.observations}"[:255] if item.observations else "Seguimiento"
                    )
                messages.info(request, f"Inspección finalizada con hallazgos. Seguimiento generado para {follow_up_date.strftime('%d/%m/%Y')}")
            else:
                inspection.status = 'Cerrada'
                inspection.save()
                
                # SI ES UN SEGUIMIENTO, verificar si el padre ya se puede cerrar definitivamente
                if inspection.parent_inspection:
                    curr = inspection
                    parent_updated = False
                    while curr.parent_inspection:
                        p = curr.parent_inspection
                        if not p.follow_ups.exclude(status__in=['Cerrada', 'Cerrada con seguimientos']).exists():
                            p.status = 'Cerrada con seguimientos'
                            p.save()
                            curr = p
                            parent_updated = True
                        else:
                            break
                    if parent_updated:
                        messages.success(request, 'Seguimiento completado y jerarquía de inspecciones cerrada.')
                    else:
                        messages.success(request, 'Seguimiento completado.')
                else:
                    if inspection.schedule_item:
                        s_item = inspection.schedule_item
                        s_item.status = 'Realizada'
                        s_item.save()
                        next_s = s_item.generate_next_schedule()
                        if next_s:
                            messages.info(request, f"📅 Nueva programación generada para {next_s.scheduled_date}")
                    messages.success(request, 'Inspección cerrada exitosamente.')
        else:
            inspection.status = 'En proceso'
            inspection.save()
            messages.success(request, 'Firma registrada. Pendiente de firmas de otros participantes.')
        
        return redirect('forklift_detail', pk=pk)

class ForkliftReportView(LoginRequiredMixin, EvidenceMixin, DetailView):
    model = ForkliftInspection
    template_name = 'inspections/forklift_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = self.object.items.all()
        context['signatures'] = self.object.signatures.select_related('user').all()
        context['total_inspected'] = items.count()
        context['total_good'] = items.filter(item_status='Bueno').count()
        context['total_bad'] = items.filter(item_status='Malo').count()
        context['any_item_has_evidence'] = any(item.evidences.exists() for item in items)
        return context

# --- REPORT MODULE VIEWS ---

class InspectionReportView(RolePermissionRequiredMixin, TemplateView):
    template_name = 'inspections/reports.html'
    permission_required = ('reports', 'view')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request
        
        # 1. Filters from GET
        f_year = request.GET.get('year', '')
        f_start = request.GET.get('start_date', '')
        f_end = request.GET.get('end_date', '')
        f_area = request.GET.get('area', '')
        f_type = request.GET.get('type', '')
        f_status = request.GET.get('status', '')
        f_responsible = request.GET.get('responsible', '')
        f_participant = request.GET.get('participant', '')

        # 2. Base Querysets & Consolidation
        # We consolidate both Scheduled items and Executed records
        
        # Model Mapping
        inspection_modules = [
            (ExtinguisherInspection, 'Extintores', 'extinguisher_detail'),
            (FirstAidInspection, 'Botiquines', 'first_aid_detail'),
            (ProcessInspection, 'Instalaciones de Proceso', 'process_detail'),
            (StorageInspection, 'Almacenamiento', 'storage_detail'),
            (ForkliftInspection, 'Montacargas', 'forklift_detail'),
        ]

        consolidated = []
        
        # Helper to apply shared filters
        def apply_filters(qs, date_field):
            if f_year:
                qs = qs.filter(**{f"{date_field}__year": f_year})
            if f_start:
                qs = qs.filter(**{f"{date_field}__gte": f_start})
            if f_end:
                qs = qs.filter(**{f"{date_field}__lte": f_end})
            if f_area:
                qs = qs.filter(area_id=f_area)
            return qs

        # A. Process Scheduled items (InspectionSchedule)
        schedule_qs = InspectionSchedule.objects.all().select_related('area', 'responsible')
        schedule_qs = apply_filters(schedule_qs, 'scheduled_date')
        if f_type:
             # Match by keyword since inspection_type is a string field
             schedule_qs = schedule_qs.filter(inspection_type__icontains=f_type)
        if f_responsible:
             schedule_qs = schedule_qs.filter(responsible_id=f_responsible)
        if f_status and f_status in ['Programada', 'Pendiente', 'Realizada']:
             schedule_qs = schedule_qs.filter(status=f_status)
        
        # For listing, we prefer actual inspection records if they exist
        # But we start with schedules to identify "Pending" ones
        scheduled_items_map = {} # To avoid duplicates in listing
        
        for item in schedule_qs:
            status_label = item.status_label
            # If item is realized, we'll pick it up from the module loop below
            if item.status == 'Realizada':
                continue
                
            consolidated.append({
                'id': f"SCH-{item.id}",
                'type': item.inspection_type,
                'area': item.area.name,
                'date_prog': item.scheduled_date,
                'date_exec': None,
                'status': status_label,
                'responsible': item.responsible.get_full_name() if item.responsible else 'N/A',
                'participants': 'N/A',
                'detail_url': None,
                'is_record': False
            })

        # B. Process Actual Inspection Records
        for model, label, detail_url_name in inspection_modules:
            if f_type and label != f_type:
                continue
                
            qs = model.objects.all().select_related('area', 'inspector')
            qs = apply_filters(qs, 'inspection_date')
            
            if f_responsible:
                qs = qs.filter(inspector_id=f_responsible)
            
            if f_status:
                # Filter by record status if provided
                # Note: labels might differ slightly between modules
                qs = qs.filter(status__icontains=f_status) if hasattr(model, 'status') else qs

            if f_participant:
                # Check if user is in signatures
                qs = qs.filter(signatures__user_id=f_participant)

            for insp in qs.distinct():
                # Get participants string
                participants = ", ".join([u.get_full_name() or u.username for u in insp.get_participants()])
                
                consolidated.append({
                    'id': f"{label[:3].upper()}-{insp.id}",
                    'type': label,
                    'area': insp.area.name,
                    'date_prog': insp.schedule_item.scheduled_date if insp.schedule_item else None,
                    'date_exec': insp.inspection_date,
                    'status': insp.status if hasattr(insp, 'status') else insp.general_status,
                    'responsible': insp.inspector.get_full_name() if insp.inspector else 'N/A',
                    'participants': participants,
                    'detail_url': reverse(detail_url_name, args=[insp.pk]),
                    'is_record': True
                })

        # 3. Aggregates for Cards
        today = date.today()
        
        # Programadas: Total items in the schedule for the filtered period
        prog_count = schedule_qs.count()
        
        # Ejecutadas: Total de registros de inspección iniciados
        ejec_count = sum(1 for x in consolidated if x['is_record'])

        # Cerradas: Solo los que están en estado Cerrada o Cerrada con Hallazgos
        closed_count = sum(1 for x in consolidated if x['is_record'] and 'Cerrada' in x['status'])
        
        # Pendientes vs Vencidas (from Scheduled items that are NOT done)
        pending_items_qs = schedule_qs.exclude(status='Realizada')
        vencidas_count = pending_items_qs.filter(scheduled_date__lt=today).count()
        pendientes_count = pending_items_qs.filter(scheduled_date__gte=today).count()
        
        context['stats'] = {
            'programmed': prog_count,
            'closed': closed_count,
            'pending': pendientes_count,
            'overdue': vencidas_count,
        }

        # 4. Filter Data for Dropdowns
        from django.contrib.auth import get_user_model
        User = get_user_model()
        context['areas'] = Area.objects.filter(is_active=True)
        context['users'] = User.objects.filter(is_active=True).order_by('first_name')
        context['types'] = [m[1] for m in inspection_modules]
        context['statuses'] = ['Programada', 'En proceso', 'Seguimiento en proceso', 'Cerrada', 'Cerrada con seguimientos']

        # Dynamic Years Data
        all_years = set()
        all_years.add(date.today().year) # Always include current year
        
        # Years from schedule
        sched_years = InspectionSchedule.objects.values_list('scheduled_date__year', flat=True).distinct()
        all_years.update([y for y in sched_years if y])
        
        # Years from modules
        for model_cls, label, _ in inspection_modules:
            mod_years = model_cls.objects.values_list('inspection_date__year', flat=True).distinct()
            all_years.update([y for y in mod_years if y])
            
        context['years_data'] = sorted(list(all_years), reverse=True)
        
        # 5. Trend Graph Data (Combined Chart)
        trend_labels = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        current_year = date.today().year
        year_to_graph = int(f_year) if f_year else current_year
        
        cerr_series = []
        pend_series = []
        venc_series = []
        
        for i in range(1, 13):
            # Cerradas: Tienen registro y el estado contiene 'Cerrada'
            cerr_count = sum(1 for x in consolidated if x['is_record'] and x['date_exec'] and x['date_exec'].month == i and x['date_exec'].year == year_to_graph and 'Cerrada' in x['status'])
            
            # Pendientes: Lo que está programado pero aún NO tiene registro real
            count_pend = sum(1 for x in consolidated if not x['is_record'] and x['date_prog'] and x['date_prog'].month == i and x['date_prog'].year == year_to_graph)
            
            # Vencidas: Programadas SIN registro cuya fecha ya pasó (mismo criterio que el contador del card)
            count_venc = sum(1 for x in consolidated if not x['is_record'] and x['date_prog'] and x['date_prog'].month == i and x['date_prog'].year == year_to_graph and x['date_prog'] < today)
            
            cerr_series.append(cerr_count)
            pend_series.append(count_pend)
            venc_series.append(count_venc)
            
        context['trend_data'] = {
            'labels': trend_labels,
            'cerradas': cerr_series,
            'pendientes': pend_series,
            'vencidas': venc_series
        }
        context['consolidated'] = sorted(consolidated, key=lambda x: (x['date_exec'] or x['date_prog']), reverse=True)
        
        return context

class InspectionReportExportView(RolePermissionRequiredMixin, View):
    permission_required = ('reports', 'view')

    def get(self, request):
        f_year = request.GET.get('year', '')
        f_start = request.GET.get('start_date', '')
        f_end = request.GET.get('end_date', '')
        f_area = request.GET.get('area', '')
        f_type = request.GET.get('type', '')
        f_status = request.GET.get('status', '')
        f_responsible = request.GET.get('responsible', '')
        f_participant = request.GET.get('participant', '')

        inspection_modules = [
            (ExtinguisherInspection, 'Extintores'),
            (FirstAidInspection, 'Botiquines'),
            (ProcessInspection, 'Instalaciones de Proceso'),
            (StorageInspection, 'Almacenamiento'),
            (ForkliftInspection, 'Montacargas'),
        ]

        def apply_filters(qs, date_field):
            if f_year: qs = qs.filter(**{f"{date_field}__year": f_year})
            if f_start: qs = qs.filter(**{f"{date_field}__gte": f_start})
            if f_end: qs = qs.filter(**{f"{date_field}__lte": f_end})
            if f_area: qs = qs.filter(area_id=f_area)
            return qs

        consolidated = []

        # Scheduled
        schedule_qs = InspectionSchedule.objects.all().select_related('area', 'responsible')
        schedule_qs = apply_filters(schedule_qs, 'scheduled_date')
        if f_type: schedule_qs = schedule_qs.filter(inspection_type__icontains=f_type)
        if f_responsible: schedule_qs = schedule_qs.filter(responsible_id=f_responsible)
        if f_status and f_status in ['Programada', 'Pendiente', 'Realizada']:
            schedule_qs = schedule_qs.filter(status=f_status)

        for item in schedule_qs:
            if item.status == 'Realizada': continue
            consolidated.append([
                f"SCH-{item.id}", item.inspection_type, item.area.name, 
                item.scheduled_date.strftime('%d/%m/%Y'), "N/A", 
                item.status_label, 
                item.responsible.get_full_name() if item.responsible else 'N/A', 
                "N/A"
            ])

        # Executed
        for model, label in inspection_modules:
            if f_type and label != f_type: continue
            qs = model.objects.all().select_related('area', 'inspector')
            qs = apply_filters(qs, 'inspection_date')
            if f_responsible: qs = qs.filter(inspector_id=f_responsible)
            if f_status: qs = qs.filter(status__icontains=f_status) if hasattr(model, 'status') else qs
            if f_participant: qs = qs.filter(signatures__user_id=f_participant)

            for insp in qs.distinct():
                participants = ", ".join([u.get_full_name() or u.username for u in insp.get_participants()])
                consolidated.append([
                    f"{label[:3].upper()}-{insp.id}", label, insp.area.name,
                    insp.schedule_item.scheduled_date.strftime('%d/%m/%Y') if insp.schedule_item else "N/A",
                    insp.inspection_date.strftime('%d/%m/%Y'),
                    insp.status if hasattr(insp, 'status') else insp.general_status,
                    insp.inspector.get_full_name() if insp.inspector else 'N/A',
                    participants
                ])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial de Inspecciones"
        
        headers = ["ID", "Tipo de Inspección", "Área", "Fecha Programada", "Fecha Ejecutada", "Estado", "Responsable", "Participantes"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="49BAA0", end_color="49BAA0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row in consolidated:
            ws.append(row)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                except: pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Reporte_Inspecciones_{date.today()}.xlsx'
        wb.save(response)
        return response
# --- Evidence Views ---
class EvidenceUploadView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        content_type_id = request.POST.get('content_type_id')
        object_id = request.POST.get('object_id')
        image = request.FILES.get('image')
        description = request.POST.get('description', '')

        if not content_type_id or not object_id or not image:
            return JsonResponse({'error': 'Faltan datos requeridos'}, status=400)

        # Basic validation
        if image.size > 5 * 1024 * 1024:
            return JsonResponse({'error': 'La imagen excede los 5MB'}, status=400)
        
        allowed_types = ['image/jpeg', 'image/png', 'image/webp']
        if image.content_type not in allowed_types:
             return JsonResponse({'error': 'Formato no permitido. Solo JPG, PNG o WEBP'}, status=400)

        try:
            content_type = ContentType.objects.get_for_id(content_type_id)
            target_obj = content_type.get_object_for_this_type(pk=object_id)
            
            # Authorization check
            inspection = None
            if hasattr(target_obj, 'status'):
                inspection = target_obj
            elif hasattr(target_obj, 'inspection'):
                inspection = target_obj.inspection
            
            if inspection and hasattr(inspection, 'status'):
                if inspection.status in ['Cerrada', 'Cerrada con seguimientos']:
                    return JsonResponse({'error': 'No se pueden agregar evidencias a una inspección cerrada'}, status=403)

            evidence = InspectionEvidence.objects.create(
                content_type=content_type,
                object_id=object_id,
                image=image,
                description=description,
                uploaded_by=request.user
            )

            return JsonResponse({
                'id': evidence.id,
                'url': evidence.image.url,
                'description': evidence.description,
                'uploaded_at': evidence.uploaded_at.strftime('%d/%m/%Y %H:%M')
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class EvidenceDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        evidence = get_object_or_404(InspectionEvidence, pk=pk)
        
        # Check status:
        # IMPORTANT: check 'inspection' FK FIRST, because item models (e.g. ExtinguisherItem)
        # also have a 'status' field ('Bueno','Malo','Recargar') which would be mistaken for
        # the inspection status if we checked hasattr(status) first.
        target_obj = evidence.content_object
        inspection = None
        if hasattr(target_obj, 'inspection'):
            # target_obj is an item (ExtinguisherItem, FirstAidItem, etc.)
            inspection = target_obj.inspection
        elif hasattr(target_obj, 'status'):
            # target_obj is the inspection itself
            inspection = target_obj

        # Block only if inspection is in a definitively closed state
        CLOSED_STATUSES = ['Cerrada', 'Cerrada con seguimientos', 'Cerrada con Hallazgos']
        if inspection and hasattr(inspection, 'status') and inspection.status in CLOSED_STATUSES:
             return JsonResponse({'error': 'No se pueden eliminar evidencias de una inspección cerrada'}, status=403)

        evidence.image.delete() # Remove file
        evidence.delete()
        return JsonResponse({'success': True})
